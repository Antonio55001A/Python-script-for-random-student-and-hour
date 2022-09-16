#Import all libraries to use
#Each library is installed by typing "pip install library name" into the command prompt as administrator

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl 
from openpyxl.chart import BarChart, Series, Reference
from random import randint
from ast import For, Pow
from cmath import sqrt
import math
from traceback import print_tb

#open the excel file
#insert the name of the excel file in the quotation marks
exceldoc = openpyxl.load_workbook("scuola1.xlsx")
sheet = exceldoc.active

excelNameSurname=openpyxl.load_workbook("nomiCognomiES.xlsx") 
sheet2=excelNameSurname.active

# I'm going to create and position the column titles

sheet["A1"].value = "Nome"
sheet["B1"].value = "Cognome"
sheet["C1"].value = "Materia1"
sheet["D1"].value = "OreMateria1"
sheet["E1"].value = "Materia2"
sheet["F1"].value = "OreMateria2"
sheet["G1"].value = "Materia3"
sheet["H1"].value = "OreMateria3"


sheet["J1"].value = "MediaMaterie"
sheet["K1"].value = "Varianza"
sheet["L1"].value = "Devianza"
sheet["M1"].value = "Covarianza"
sheet["N1"].value = "Coefficiente Angolare"


#here I go to get the first and last names in the second file I uploaded
#in n I'm going to enter the number of how many students I want to generate
#the function will fetch the names and surnames from the excel file and automatically insert them in the new file

def GeneraStudenti(n):
    d=0
    e=0
    i=0
    matriceStudenti = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 1).value = sheet2.cell(row=i+3,column=1).value
        d=sheet.cell(row = i + 2, column= 2).value = sheet2.cell(row=i+3,column=2).value
        matriceStudenti.append(np.array([c, d]))
        i+=1
    return matriceStudenti



# 1) Thanks to this function you put values in tables
# 2) Thanks to AL FOR, I go to insert 3 elements horizontally at each turn, 1 for each subject
# 3) Then save the ary of the elements placed in an array
# 4) Once this is done, I repeat the tour until I finish the lines I want to put
# 5) The number you pass will indicate how many lines it will generate


def GeneraVoti(n):
    d=0
    e=0
    i=0
    matriceVoti = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 3).value = randint(1, 30)
        d=sheet.cell(row = i + 2, column= 5).value = randint(1, 30)
        e=sheet.cell(row = i + 2, column= 7).value = randint(1, 30)
        matriceVoti.append(np.array([c, d, e]))
        i+=1
    return matriceVoti

#does the same thing as the vote function

def GeneraOre(n):
    d=0
    e=0
    i=0
    matriceOre = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 4).value = randint(1, 15)
        d=sheet.cell(row = i + 2, column= 6).value = randint(1, 15)
        e=sheet.cell(row = i + 2, column= 8).value = randint(1, 15)
        matriceOre.append(np.array([c, d, e]))
        i+=1
    return matriceOre



# 1) an array is passed as a parameter
# 2) the while loop is for fetching rows of items by row
# 3) once the row is taken it is saved in a temporary variable and the sum of element by element of the single row is made
# 4) once this is done, the for loop is exited and the numbers in the row are averaged
# 5) later it is placed in the excel file
# 6) is saved inside an array
# 7) start over with the next line
# 8) returns a list with all the averages of the elements in the file

def calcolaMediaMatrice(matriceTrasposta,n):
    somma=0
    i=0
    listaMedie=[]

    while i<n:
        linea=[row[i] for row in matriceTrasposta]
        for c in linea:
            somma+=c
        media=somma/len(linea)
        sheet.cell(row = i + 2, column= 10).value = media
        listaMedie.append(media)
        i+=1
    return listaMedie



# 1) here we pass the transposed matrix with all sorted elements and the list of all averages
# 2) later through the while loop I go to fetch the single row inside the array in the matrix
# 3) once the list of elements is saved I execute a for to take each element of the single row
# 4) in the for I am going to execute the formula to have the squared deviation of each element
# 5) once the difference has been obtained, it is added up to have the sum of all the differences in the square
# 6) I calculate the variance and place it in its cell
# 7) I then calculate the deviance of the variance and the position
# 8) each single deviance is put into a dedicated array
# 9) restart the while loop until the arrays inside the array are finished
# 10) return a list with all the deviances


def CalcolaDevianza(matriceTrasposta,medie,n):
    i=0
    somma=0
    b=0
    listaVarianze=[]


# cycle to calculate sum of rejects
    while i<n:

        colonna=[row[i] for row in matriceTrasposta]
        print("sono la colonna: ")
        print(colonna)
        print("sono la media della colonna")
        print(medie)
        for c in colonna:
        # here the offset is calculated and then squared
            scartoAlQuadrato=math.pow(c-medie[b],2)
            print("sono la scarto: ")
            print(scartoAlQuadrato)
            #at this point, the sum of all the discards is added
            somma+=scartoAlQuadrato
        b+=1
        #b is incremented every time a group of elements ends. In this way the 2nd group will calculate the average with that of the 2nd group
         # calculate variance by adding up the offsets / number of elements
        varianza=somma/len(colonna)
        sheet.cell(row = i + 2, column= 11).value = varianza
        # calculate the deviance by taking the square root of the variance
        devianza=math.sqrt(varianza)
        sheet.cell(row = i + 2, column= 12).value = devianza
       # returns the deviance
        listaVarianze.append(varianza)
        i+=1

    return listaVarianze


# 1) To the function I pass a matrix of grades (group A) and a matrix of hours (Group B),
# 2) I also pass the counter and the average of group A and the average of group B
# 3) I assign the variables of the function
# 4) Inside the while loop I'm going to get and save column [i] of group A and column [i] of group b
# 5) once taken I go to calculate the two elements by multiplication
# 6) later I go to run the product
# 7) once the product is done I go to save it and do the sum using the variable sum
# 8) then continues to spin in the inner loop until the elements of the single group run out
# 9) when finished it adds up divided by the number of elements
# 10) the elements of each group are taken and subtracted with the group's own average
# 11) when the elements of the group run out, the loop is exited and the average is increased and the counter b returns to 0 to iterate the elements of the next group
# 12) I am going to put the result on excel and save it in an array in which the individual covariances of each group will go
# 13) at the end I return a list with all the single covariances of each group


def CalcolaCovarianza(matriceVoti,matriceOre,n,listaMedieVoti,listeMedieOre):

    i=0
    somma=0
    b=0
    listaCovarianza=[]


    while i<n:

        colonnaVoti=[row[i] for row in matriceVoti]
        colonnaOre=[row[i] for row in matriceOre]



        while b<len(colonnaOre):

            elemento1=colonnaVoti[b]-listaMedieVoti[i]
            elemento2=colonnaOre[b]-listeMedieOre[i]
            #The elements of each group are taken and subtracted with the group's own average
            #when the elements of the group run out, the loop is exited and the average is increased and the counter b returns to 0 to iterate the elements of the next group

            prodotto=elemento1*elemento2
            somma+=prodotto
            b+=1


        covarianza=somma/b
        sheet.cell(row = i + 2, column= 13).value = covarianza
        listaCovarianza.append(covarianza)
        i+=1
        somma=0
        b=0

    return listaCovarianza



# 1) I pass the list of covariances and the list of variances to the function
# 2) in the while loop I go to calculate the coefficient of the single group
# 3) once calculated I put it on excel and save it in a list
# 4) the loop continues until the elements run out
# 5) returns a list with all the coefficients of each single group


def CoefficienteAngolare(listaCovarianze,listaVarianza,n):
    i=0
    listaCoefficinti=[]
    while i<n:
        coefficienteAngolare=listaCovarianze[i]/listaVarianza[i]
        sheet.cell(row = i + 2, column= 14).value = coefficienteAngolare
        listaCoefficinti.append(coefficienteAngolare)
        i+=1
    return listaCoefficinti


#-----------------------------------------------------------
#From here the whole file is generated
# ---------------------------------------------------------- 

# I pass the number of students for which I want to generate data

studenti=10

GeneraStudenti(studenti)

# I generate votes
matrice=GeneraVoti(studenti)
print("sono la matrice standard: ")
print(matrice)
print("la lunghezza della matrice è: ")
print(len(matrice))

# I generate the hours for each subject
matriceOre=GeneraOre(studenti)
print("le ore degli studenti sono: ")


# transformed the normal hour matrix into a transposed matrix
matriceTraspostaORE=np.transpose(matriceOre)
print(matriceTraspostaORE)


## transformed the normal matrix of votes into a transposed matrix
matriceTrasposta=np.transpose(matrice)
print("sono la matrice Trasposta: ")
print(matriceTrasposta)


#calculating the average of the marks
medieVoti=calcolaMediaMatrice(matriceTrasposta,studenti)
print("sono la media: ")
print(medieVoti)

#calculating the average of the hours
medieOre=calcolaMediaMatrice(matriceTraspostaORE,studenti)
print("sono la media: ")
print(medieOre)


#calculation the variance is the deviance 
varianza=CalcolaDevianza(matriceTrasposta,medieVoti,studenti)
print("sono la varianza: ")
print(varianza)


#calculating the covariance
covarianza=CalcolaCovarianza(matriceTrasposta,matriceTraspostaORE,studenti,medieVoti,medieOre)
print("sono la covarianza: ")
print(covarianza)

#calculate the angular coefficient
Coefficiente=CoefficienteAngolare(covarianza,varianza,studenti)
print("sono il coefficiente: ")
print(Coefficiente)


#all data is saved on excel through functions

# I'm going to save the file with the new data 
exceldoc.save("scuola5.xlsx")