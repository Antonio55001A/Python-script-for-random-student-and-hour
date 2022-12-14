#Importo tutte le librerie da utilizzare
#Le librerie si installano scrivendo "pip install nomeLibreria" 

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl 
from openpyxl.chart import BarChart, Series, Reference
from random import randint
from ast import For, Pow
from cmath import sqrt
import math
from traceback import print_tb

#apro i file excel
#inserire il nome del file excel tra le virgolette

exceldoc = openpyxl.load_workbook("scuola1.xlsx")
sheet = exceldoc.active

excelNomiCognomi=openpyxl.load_workbook("nomiCognomiES.xlsx") 
sheet2=excelNomiCognomi.active


# vado a creare e posizionare i titoli delle colonne 

sheet["A1"].value = "Nome"
sheet["B1"].value = "Cognome"
sheet["C1"].value = "Materia1"
sheet["D1"].value = "OreMateria1"
sheet["E1"].value = "Materia2"
sheet["F1"].value = "OreMateria2"
sheet["G1"].value = "Materia3"
sheet["H1"].value = "OreMateria3"


sheet["J1"].value = "MediaMaterie"
sheet["K1"].value = "Varianza"
sheet["L1"].value = "Devianza"
sheet["M1"].value = "Covarianza"
sheet["N1"].value = "Coefficiente Angolare"


#qui vado a prendere i nomi e i cognomi nel secondo file che ho caricato
#in n vado ad inserire il numero di quanti studenti voglio generare
#la funzione andrà a prendere dal file excel i nomi e i cognomi è li andrà ad inserire in automatico nel nuovo file

def GeneraStudenti(n):
    d=0
    e=0
    i=0
    matriceStudenti = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 1).value = sheet2.cell(row=i+3,column=1).value
        d=sheet.cell(row = i + 2, column= 2).value = sheet2.cell(row=i+3,column=2).value
        matriceStudenti.append(np.array([c, d]))
        i+=1
    return matriceStudenti


#1) Grazie a questa funzione metti i valori nelle tabelle 
#2) Grazie AL FOR vado a fare ad ogni giro l'inserimento di 3 elementi in maniera orizzontale, 1 per ogni materia 
#3) In seguito salva l'arrey degli elementi messi in una matrice 
#4) Fatto ciò ricomancii il giro fino a che non termino le righe che voglio mettere
#5) Il numero che si passa indicherà quante righe genererà


def GeneraVoti(n):
    d=0
    e=0
    i=0
    matriceVoti = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 3).value = randint(1, 30)
        d=sheet.cell(row = i + 2, column= 5).value = randint(1, 30)
        e=sheet.cell(row = i + 2, column= 7).value = randint(1, 30)
        matriceVoti.append(np.array([c, d, e]))
        i+=1
    return matriceVoti

#fa la stessa cosa della funzione dei voti 

def GeneraOre(n):
    d=0
    e=0
    i=0
    matriceOre = []
    while i<n:
        c=sheet.cell(row = i + 2, column= 4).value = randint(1, 15)
        d=sheet.cell(row = i + 2, column= 6).value = randint(1, 15)
        e=sheet.cell(row = i + 2, column= 8).value = randint(1, 15)
        matriceOre.append(np.array([c, d, e]))
        i+=1
    return matriceOre



#1) viene passato come parametro una matrice
#2) il ciclo while serve per andare a prendere righa di elementi per riga
#3) una volta presa la riga viene salavata in una variabile temporanea e viene fatta la somma di elemento per elemento della singola riga
#4) una volta fatto ciò si esce dal ciclo for e viene fatta la media dei numeri della riga
#5) in seguito viene posizionata nel file excel 
#6) viene salvata all'interno di un array
#7) ricoomincia con la riga successiva 
#8) ritorna una lista con tutte le medie degli elementi nel file 

def calcolaMediaMatrice(matriceTrasposta,n):
    somma=0
    i=0
    listaMedie=[]
    #passo la matrice standard poichè quella transposta ha una lunghezza delle colonne . Matrice =20, matrice trasposta=3 lunghezza
    while i<n:
        linea=[row[i] for row in matriceTrasposta]
        for c in linea:
            somma+=c
        media=somma/len(linea)
        sheet.cell(row = i + 2, column= 10).value = media
        listaMedie.append(media)
        i+=1
    return listaMedie



#1) qui passiamo la matrice trasposta con tutti gli elementi ordinati e le la lista di tutte le medie 
#2) in seguito tramite il ciclo while vado a prendere la singola riga all'interno dell'array nella matrice 
#3) una vola salvata la lista di elementi eseguo un for per prendere ogni elemento della singola riga 
#4) nel for vado ad eseguire la formula per avere lo scarto a quadrato di ogni elemento
#5) una volta avuto lo scarto questo viene sommato fino ad avere la somma di tutti gli scarti al qudrato 
#6) calcolo la varianza e la posiziono nella sua cella 
#7) in seguito calcolo la devianza della varianza e la posizione 
#8) ogni singola devianza viene messa in un array apposta 
#9) ricomincia il ciclo while fino a quando non sono finiti gli array all'interno della matrice 
#10) ritorna una lista con tutte le devianze


#passo la matrice standard poichè quella transposta ha una lunghezza delle colonne . Matrice =20, matrice trasposta=3 lunghezza


def CalcolaDevianza(matriceTrasposta,medie,n):
    i=0
    somma=0
    b=0
    listaVarianze=[]


    # cliclo per calcolare somma degli scarti
    while i<n:

        colonna=[row[i] for row in matriceTrasposta]
        print("sono la colonna: ")
        print(colonna)
        print("sono la media della colonna")
        print(medie)
        for c in colonna:
        # qui viene calcolato lo scarto e poi viene elavato al quadrato
            scartoAlQuadrato=math.pow(c-medie[b],2)
            print("sono la scarto: ")
            print(scartoAlQuadrato)
            #in questo punto si fa la somma di tutti gli scarti 
            somma+=scartoAlQuadrato
        b+=1
        #b viene incrementato ogni volta che termina un gruppo di elementi. In questo modo il 2 gruppo calcolerà la media con quella del 2 gruppo
        # calcola la varianza facendo la somma degli scarti / il numeri di elementi
        varianza=somma/len(colonna)
        sheet.cell(row = i + 2, column= 11).value = varianza
        #calcola la devianza facendo la radice quadrata della varianza
        devianza=math.sqrt(varianza)
        sheet.cell(row = i + 2, column= 12).value = devianza
        #ritorna la devianza 
        listaVarianze.append(varianza)
        i+=1

    return listaVarianze


#1) Alla funzione passo una mtrice di voti(gruppo A) e una matrice di ore(Gruppo B), 
#2) passo anche il contatore e la media del gruppo A e la media del gruppo B
#3) assegno le variabili della funzione 
#4) All'interno del ciclo while vado a prendere e salavare La colonna[i] del gruppo A e la colonna[i] del gruppo b
#5) una volta presa vado a calcolare i due elementi per la moltiplicazione
#6) in seguito vado ad eseguire il prodotto 
#7) una volta fatto il prodotto vado a salavarlo e a fare la somma tramite la variabile somma
#8) in seguito continua a girare nel ciclo interno fin quando non finiscono gli elementi del singolo gruppo
#9) una volta terminati fa la somma diviso il numero di elementi 
#10) vengono presi gli elementi di ogni gruppo e sottratti con la propia media del gruppo
#11) quando finiscono gli elementi del gruppo si esce dal ciclo e la media viene aumentata e il counter b ritorna a 0 per iterare gli elementi del prossimo gruppo
#12) il risultato lo vado a mettere su excel e a salvare all'interno di un'array nella quale andranno le singole covarianze di ogni gruppo 
#13) alla fine ritorno una lista con tutte le singole covarianze di ogni gruppo 


def CalcolaCovarianza(matriceVoti,matriceOre,n,listaMedieVoti,listeMedieOre):

    i=0
    somma=0
    b=0
    listaCovarianza=[]


    while i<n:

        colonnaVoti=[row[i] for row in matriceVoti]
        colonnaOre=[row[i] for row in matriceOre]



        while b<len(colonnaOre):

            elemento1=colonnaVoti[b]-listaMedieVoti[i]
            elemento2=colonnaOre[b]-listeMedieOre[i]
            #vengono presi gli elementi di ogni gruppo e sottratti con la propia media del gruppo
            #quando finiscono gli elementi del gruppo si esce dal ciclo e la media viene aumentata e il counter b ritorna a 0 per iterare gli elementi del prossimo gruppo

            prodotto=elemento1*elemento2
            somma+=prodotto
            b+=1


        covarianza=somma/b
        sheet.cell(row = i + 2, column= 13).value = covarianza
        listaCovarianza.append(covarianza)
        i+=1
        somma=0
        b=0

    return listaCovarianza



#1) alla funzione passo la lista delle covarianze e la lista delle varianze 
#2) nel ciclo while vado a calcolare il coefficiente del singolo gruppo 
#3) una volta calcolato lo metto su excel e lo salvo in una lista 
#4) il ciclo continua finchè non finiscono gli elementi
#5) ritorna una lista con tutti i coefficienti di ogni singolo gruppo


def CoefficienteAngolare(listaCovarianze,listaVarianza,n):
    i=0
    listaCoefficinti=[]
    while i<n:
        coefficienteAngolare=listaCovarianze[i]/listaVarianza[i]
        sheet.cell(row = i + 2, column= 14).value = coefficienteAngolare
        listaCoefficinti.append(coefficienteAngolare)
        i+=1
    return listaCoefficinti


#-----------------------------------------------------------
#Da qui viene generato tutto il file
# ---------------------------------------------------------- 

#passo il numero di studenti per la quale voglio generare i dati 
studenti=90


GeneraStudenti(studenti)

#genero i voti 
matrice=GeneraVoti(studenti)
print("sono la matrice standard: ")
print(matrice)
print("la lunghezza della matrice è: ")
print(len(matrice))

#genero le ore per singola materia 
matriceOre=GeneraOre(studenti)
print("le ore degli studenti sono: ")


#trasformo la matrice normale delle ore in matrice transposta
matriceTraspostaORE=np.transpose(matriceOre)
print(matriceTraspostaORE)


##trasformo la matrice normale dei voti in matrice transposta
matriceTrasposta=np.transpose(matrice)
print("sono la matrice Trasposta: ")
print(matriceTrasposta)


#calcolo la media dei voti 
medieVoti=calcolaMediaMatrice(matriceTrasposta,studenti)
print("sono la media: ")
print(medieVoti)

#calcolo la media delle ore
medieOre=calcolaMediaMatrice(matriceTraspostaORE,studenti)
print("sono la media: ")
print(medieOre)


#calcolo la varianza e la devianza 
varianza=CalcolaDevianza(matriceTrasposta,medieVoti,studenti)
print("sono la varianza: ")
print(varianza)


#calcolo la covarianza
covarianza=CalcolaCovarianza(matriceTrasposta,matriceTraspostaORE,studenti,medieVoti,medieOre)
print("sono la covarianza: ")
print(covarianza)

#calcolo il coefficiente anngolare
Coefficiente=CoefficienteAngolare(covarianza,varianza,studenti)
print("sono il coefficiente: ")
print(Coefficiente)


#tutti i dati vengono salavati su excel tramite le funzioni 

#vado a salvare il file con i nuovi dati 
exceldoc.save("scuola5.xlsx")